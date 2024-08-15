from openpyxl import load_workbook
from components.exportErrors import load6312Error, load6320Error, loadGateway, load6312GatewayErr
from datetime import datetime
from components.loadRawData import raw6312, raw6320, raw6303, raw63123
from components.readXML import load6312XML, load6320XML, load6303XML
from components.certNumb import certificateNumber
from components.gatewayConfig import gatewayJson
import os

def cert6312(voltage, product, spec, customer, address, regNum, meterTest, xmlFile, csvFile, badgeNum, firmware, PTnum,):
    
    openEmptyCert = load_workbook(r"excelFiles\emptyCertificate.xlsx")
    sheetCert = openEmptyCert.active

    
    if sheetCert.cell(row=10, column=4).value == "RX-30-430Xytronic":
        bench = "WECO4150"

    elif sheetCert.cell(row=10, column=4).value == "RM-20-102": 
        bench = "WECO2350"

    if spec == '6312 (1-12)' or spec == '6312 (13-24)':
        seriesData, voltageData = load6312GatewayErr(voltage, spec)
    
    else:
        seriesData, voltageData, voltageData3, seriesData3 = load6312Error(voltage, bench)


    if product == '6312-3P':
            (
        sequenceValue,
        meterID,
        serialNum,
        modelType,
        leftCoil,
        rightCoil,
        PFValue,
        FLValue,
        LLValue,
        middleCoil
        ) = raw63123(csvFile)

    else:
        (
        sequenceValue,
        meterID,
        serialNum,
        modelType,
        leftCoil,
        rightCoil,
        PFValue,
        FLValue,
        LLValue,
        ) = raw6312(csvFile)

    macAddress, modelNum, unitSerial, meterConfig, demandInt, pulseType = load6312XML(xmlFile)

    if str(serialNum[0]) not in unitSerial:
        return print("Serial Numbers don't match")

    newCertnum = certificateNumber()

    certificate = load_workbook(r"excelFiles\emptyCertificate.xlsx")
    
    if product == '6312-3P':
        sheet = certificate["6312-3P"]
    else:
        sheet = certificate["6312"]

    # Certificate Number
    sheet.cell(row=2, column=4).value = f"{newCertnum}"

    # Owner/Customer, Address and Seal Expiry
    if customer == "Metergy":
        sheet.cell(row=3, column=4).value = "Metergy Solutions Inc."
        sheet.cell(
            row=3, column=16
        ).value = "Suite 601, 8133 Warden Ave., Markham, ON, Canada, L6G 1B3"
        sheet.cell(row=2, column=25).value = "90902"

    else:
        sheet.cell(row=3, column=4).value = customer
        sheet.cell(row=3, column=16).value = address
        sheet.cell(row=2, column=25).value = regNum


    # Verification & Seal Expiry Date

    todayDate = datetime.now()
    currentYear = int(todayDate.year)

    if meterTest == "Verified":
        sheet.cell(row=22, column=4).value = "Verified"
        newYear = currentYear + 10
        sheet.cell(row=2, column=16).value = newYear

    elif meterTest == "Re-verified":
        sheet.cell(row=22, column=4).value = "Re-Verified"
        newYear = currentYear + 8
        sheet.cell(row=2, column=16).value = newYear

    # MAC Address, Serial Number, Model Number, Pulse Weight, Register Type (Display)

    sheet.cell(row=7, column=23).value = macAddress

    sheet.cell(row=5, column=16).value = unitSerial
    sheet.cell(row=5, column=21).value = unitSerial

    
    if product == '6312-3P':
        for row in range(26, 34):
            sheet.cell(row=row, column=4).value = unitSerial
    else:
        for row in range(26, 38):
            sheet.cell(row=row, column=4).value = unitSerial

    sheet.cell(row=7, column=4).value = modelNum

    sheet.cell(row=8, column=9).value = "Display"

    if voltage == "347":
        sheet.cell(row=8, column=4).value = "0.0012"
    elif voltage == "480":
        sheet.cell(row=8, column=4).value = "0.0016"
    elif voltage == "600":
        sheet.cell(row=8, column=4).value = "0.002"
    else:
        sheet.cell(row=8, column=4).value = "0.0004"

    # Output Pulse
    sheet.cell(row=8, column=23).value = "N/A"

    # Element, Current, Voltage
    sheet.cell(row=6, column=21).value = voltage + "V"
    sheet.cell(row=6, column=25).value = "0.80-80mA"

    # Badge Number
    sheet.cell(row=5, column=4).value = badgeNum + "-01"
    
    if product == '6312-3P':
        sheet.cell(row=5, column=8).value = badgeNum + "-8"
    else:
        sheet.cell(row=5, column=8).value = badgeNum + "-12"

    startNum = 1

    if product == '6312-3P':
        for row in range(26, 34):
            newBadge = f"{badgeNum}-{startNum:02}"
            sheet.cell(row=row, column=3).value = newBadge
            startNum += 1
    
    else: 
        for row in range(26, 38):
            newBadge = f"{badgeNum}-{startNum:02}"
            sheet.cell(row=row, column=3).value = newBadge
            startNum += 1

    # Section E / Test Passing

    sheet.cell(row=22, column=9).value = "Passed"
    sheet.cell(row=22, column=25).value = "Passed"
    sheet.cell(row=23, column=4).value = "Passed"
    sheet.cell(row=23, column=25).value = "Passed"

    # Error Corrections
    sheet.cell(row=17, column=5).value = "Y"
    sheet.cell(row=17, column=6).value = "Y"
    sheet.cell(row=17, column=7).value = "Y"
    sheet.cell(row=17, column=9).value = "Y"
    sheet.cell(row=17, column=19).value = "Y"

    if product == '6312-3P':
        sheet.cell(row=17, column=8).value = "Y"

    # Config and Element
    if product == "6312-1P":
        sheet.cell(row=7, column=14).value = "1P-12 3W 1PH (2EL)"
        sheet.cell(row=6, column=16).value = 2
    elif product == "6312-2P":
        sheet.cell(row=7, column=14).value = "2P-12 3W Network (2EL)"
        sheet.cell(row=6, column=16).value = 2
    elif product == "6312-3P":
        sheet.cell(row=7, column=14).value = "3P-08 4W Y (3EL)"
        sheet.cell(row=6, column=16).value = 3

    if spec == '6312 (1-12)' or spec == '6312 (13-24)':
        sheet.cell(row=7, column=14).value = "3W Network (2EL)"
        sheet.cell(row=6, column=16).value = 2

    # Firmware Version

    sheet.cell(row=8, column=16).value = firmware

    # PT SN and Ratio 

    if voltage == "347":
        sheet.cell(row=19, column=16).value = "PT Module"
        sheet.cell(row=19, column=18).value = "S/N"
        sheet.cell(row=19, column=19).value = PTnum

        sheet.cell(row=19, column=21).value = "347V 60Hz / WYE, PT Ratio 2.892:1"

    elif voltage == "600":
        sheet.cell(row=19, column=16).value = "PT Module"
        sheet.cell(row=19, column=18).value = "S/N"
        sheet.cell(row=19, column=19).value = PTnum

        sheet.cell(row=19, column=21).value = "600V 60Hz / DELTA, PT Ratio 5:1"

    
    # Test JIG Serials (Verify)
    if bench == "WECO2350":
        if spec == "6312 (1-12)" or spec == "6312 (13-24)":
            sheet.cell(row=12, column=4).value = "TRIACTA-GT-XP-01"
            sheet.cell(row=12, column=10).value = "MSI-9150JIG-01"
        
        else:
            sheet.cell(row=12, column=4).value = "TRIACTA-6312-JIG-XP-01"
            sheet.cell(row=12, column=10).value = "ECI-9120JIG-01"
        

    # Calculations for bottom of certificate

    newFL = []
    newLL = []
    newPF = []
    newLeft = []
    newRight = []
    newMiddle = []
    averageValues = []

    if product == "6312-2P" or product == "6312-1P" or spec == "6312 (1-12)" or spec == "6312 (13-24)":
        seriesValues = seriesData[voltage]
        voltageValues = voltageData[voltage]

    elif product == "6312-3P":
        seriesValues = seriesData3[voltage]
        voltageValues = voltageData3[voltage]

    if product == "6312-2P" or product == "6312-1P" or spec == "6312 (1-12)" or spec == "6312 (13-24)":

        for i in seriesValues:
            newValueFL = round(FLValue[seriesValues.index(i)] - i[1], 2)
            newFL.append(newValueFL)

            newValueLL = round(LLValue[seriesValues.index(i)] - i[0], 2)
            newLL.append(newValueLL)

            newValuePF = round(PFValue[seriesValues.index(i)] - i[2], 2)
            newPF.append(newValuePF)

        for index, i in enumerate(voltageValues):
            if index % 2 == 0:
                newLeftValue = round(leftCoil[index // 2] - i[2], 2)
                newLeft.append(newLeftValue)
            else:
                newRightValue = round(rightCoil[index // 2] - i[2], 2)
                newRight.append(newRightValue)

        for index, row in enumerate(range(26, 38)):
            sheet.cell(row=row, column=5).value = newFL[index]
            sheet.cell(row=row, column=13).value = newLL[index]
            sheet.cell(row=row, column=9).value = newPF[index]
            sheet.cell(row=row, column=10).value = newLeft[index]
            sheet.cell(row=row, column=12).value = newRight[index]

            average = round(
            (
                    abs(newFL[index])
                    + abs(newLL[index])
                    + abs(newPF[index])
                    + abs(newLeft[index])
                    + abs(newRight[index])
                )
                / 5,
                2,
            )
            averageValues.append(average)
            sheet.cell(row=row, column=23).value = averageValues[index]

    elif product == "6312-3P":

        for index, i in enumerate(seriesValues):
            newValueFL = round(FLValue[index] - i[1], 2)
            newFL.append(newValueFL)

            newValueLL = round(LLValue[index] - i[0], 2)
            newLL.append(newValueLL)


            newValuePF = round(PFValue[index] - i[2], 2)
            newPF.append(newValuePF) 

        for index, i in enumerate(voltageValues):
            if index % 3 == 0:
                newLeftValue = round(leftCoil[index // 3] - i[2], 2)
                newLeft.append(newLeftValue)

            elif index % 3 == 1:
                newMiddleValue = round(middleCoil[index // 3] - i[2], 2)
                newMiddle.append(newMiddleValue)

            else:
                newRightValue = round(rightCoil[index // 3] - i[2], 2)
                newRight.append(newRightValue)

        for index, row in enumerate(range(26, 34)):
            sheet.cell(row=row, column=5).value = newFL[index]
            sheet.cell(row=row, column=13).value = newLL[index]
            sheet.cell(row=row, column=9).value = newPF[index]
            sheet.cell(row=row, column=10).value = newLeft[index]
            sheet.cell(row=row, column=12).value = newRight[index]
            sheet.cell(row=row, column=11).value = newMiddle[index]

            average = round(
            (
                    abs(newFL[index])
                    + abs(newLL[index])
                    + abs(newPF[index])
                    + abs(newLeft[index])
                    + abs(newRight[index])
                    + abs(newMiddle[index])
                )
                / 6,
                2,
            )
            averageValues.append(average)
            sheet.cell(row=row, column=23).value = averageValues[index]

    # Name for the worksheet must be dynamic and alternating
    # Delete other sheets later
    if product == '6312-3P':
        certificate.remove(certificate["6312"])
    elif product == '6312-1P' or product == '6312-2P' or spec == "6312 (1-12)" or spec == "6312 (13-24)":
        certificate.remove(certificate["6312-3P"])
    certificate.remove(certificate["6320"])
    certificate.remove(certificate["6303"])
    certificate.remove(certificate["Gateway1"])
    certificate.remove(certificate["Gateway2"])

    filePath = r"outputs/modifiedCert.xlsx"
 
    if os.path.exists(filePath):
        print(f"The file '{filePath}' already exists.")
    else:
    # If the file doesn't exist, save the certificate as a new file
        os.makedirs(os.path.dirname(filePath), exist_ok=True)  # Create directory if it doesn't exist
        print(filePath)
        certificate.save(filePath)
        print(f"The file '{filePath}' has been created and saved.")
 
    # If you want to ensure the workbook is saved regardless:
    certificate.save(filePath)

    openEmptyCert.close()
    certificate.close()


def cert6320(type, customer, address, regNum, meterTest, xmlFile, csvFile, badgeNum, firmware):
    
    openEmptyCert = load_workbook(r"excelFiles\emptyCertificate.xlsx")
    sheetCert = openEmptyCert.active

    
    if sheetCert.cell(row=10, column=4).value == "RX-30-430Xytronic":
        bench = "WECO4150"

    elif sheetCert.cell(row=10, column=4).value == "RM-20-102": 
        bench = "WECO2350"
   
    seriesData, voltageData = load6320Error("120", type, bench)

    (
        sequenceValue,
        meterID,
        serialNum,
        modelType,
        leftCoil,
        rightCoil,
        PFValue,
        FLValue,
        LLValue,
        SeriesValue,
    ) = raw6320(csvFile)

    macAddress, modelNum, unitSerial = load6320XML(xmlFile)

    newCertnum = certificateNumber()

    if str(serialNum[0]) not in unitSerial:
        return print("Serial Numbers don't match")

    certificate = load_workbook(r"excelFiles\emptyCertificate.xlsx")
    sheet = certificate["6320"]

    # Certificate Number
    sheet.cell(row=2, column=4).value = f"{newCertnum}"

    # Owner/Customer, Address and Seal Expiry
    if customer == "Metergy":
        sheet.cell(row=3, column=4).value = "Metergy Solutions Inc."
        sheet.cell(
            row=3, column=16
        ).value = "Suite 601, 8133 Warden Ave., Markham, ON, Canada, L6G 1B3"
        sheet.cell(row=2, column=25).value = "90902"

    else:
        sheet.cell(row=3, column=4).value = customer
        sheet.cell(row=3, column=16).value = address
        sheet.cell(row=2, column=25).value = regNum


    # Verification & Seal Expiry Date

    todayDate = datetime.now()
    currentYear = int(todayDate.year)

    if meterTest == "Verified":
        sheet.cell(row=22, column=4).value = "Verified"
        newYear = currentYear + 10
        sheet.cell(row=2, column=16).value = newYear

    elif meterTest == "Re-verified":
        sheet.cell(row=22, column=4).value = "Re-Verified"
        newYear = currentYear + 8
        sheet.cell(row=2, column=16).value = newYear

    # MAC Address, Serial Number, Model Number, Pulse Weight, Register Type (Display)

    sheet.cell(row=7, column=23).value = macAddress

    sheet.cell(row=5, column=16).value = unitSerial
    sheet.cell(row=5, column=21).value = unitSerial

    for row in range(27, 47):
        sheet.cell(row=row, column=4).value = unitSerial

    sheet.cell(row=7, column=4).value = modelNum

    sheet.cell(row=8, column=9).value = "Display"

    sheet.cell(row=8, column=4).value = "0.0004"

    # Output Pulse
    sheet.cell(row=8, column=23).value = "N/A"

    # Element, Current, Voltage
    sheet.cell(row=6, column=21).value = "120V"
    sheet.cell(row=6, column=25).value = "0.80-80mA"

    # Badge Number
    sheet.cell(row=5, column=4).value = badgeNum + "-01"
    sheet.cell(row=5, column=8).value = badgeNum + "-12"

    startNum = 1
    for row in range(27, 47):
        newBadge = f"{badgeNum}-{startNum:02}"
        sheet.cell(row=row, column=3).value = newBadge
        startNum += 1

    # Section E / Test Passing

    sheet.cell(row=22, column=9).value = "Passed"
    sheet.cell(row=22, column=25).value = "Passed"
    sheet.cell(row=23, column=4).value = "Passed"
    sheet.cell(row=23, column=25).value = "Passed"


    # Error Corrections
    sheet.cell(row=17, column=5).value = "Y"
    sheet.cell(row=17, column=6).value = "Y"
    sheet.cell(row=17, column=7).value = "Y"
    sheet.cell(row=17, column=9).value = "Y"
    sheet.cell(row=17, column=19).value = "Y"
    sheet.cell(row=17, column=20).value = "Y"

    # Config and Element
    if type == "6320-3P":
        sheet.cell(row=7, column=14).value = "3P-10 4W Y (3EL) and 10"
        sheet.cell(row=6, column=16).value = 3

    elif type == "6320-2P":
        sheet.cell(row=7, column=14).value = "2P-20 3W 1PH Network (2EL)"
        sheet.cell(row=6, column=16).value = 2

    elif type == "6320-1P":
        sheet.cell(row=7, column=14).value = "1P-20 3W 1PH (2EL)"
        sheet.cell(row=6, column=16).value = 2

    # Firmware Version

    sheet.cell(row=8, column=16).value = firmware

    # Test Jig 

    if bench == "WECO2350":
        sheet.cell(row=12, column=4).value = "TRIACTA-6320-JIG-XP-01"
        sheet.cell(row=12, column=10).value = "ECI-9120JIG-03"

    # Calculations for bottom of certificate

    newFL = []
    newLL = []
    newPF = []
    newLeft = []
    newRight = []
    newPPF = []
    averageValues = []

    seriesValues = seriesData[type]
    voltageValues = voltageData["120"]

    # FL, LL, PF , Left and Right values

    for index, i in enumerate(seriesValues):
        newValueFL = round(FLValue[index] - i[1], 3)
        newFL.append(newValueFL)

        newValueLL = round(LLValue[index] - i[0], 3)
        newLL.append(newValueLL)

        newValuePF = round(PFValue[seriesValues.index(i)] - i[2], 3)
        newPF.append(newValuePF)

        newValuePPF = round(SeriesValue[seriesValues.index(i)] - i[3], 3)
        newPPF.append(newValuePPF)

    # Right and Left coil calculations (Iterates)

    for index, i in enumerate(voltageValues[:15]):
        if index % 3 == 0:
            newLeftValue = round(leftCoil[index // 3] - i[2], 3)
            newLeft.append(newLeftValue)
        if index >= 2 and (index + 1) % 3 == 0:
            newRightValue = round(rightCoil[(index + 1) // 3 - 1] - i[2], 3)
            newRight.append(newRightValue)

    for index, i in enumerate(voltageValues[15:25]):
        actualIndex = index + 15

        if actualIndex % 2 != 0:
            coilIndex = actualIndex // 2 - 2
            newLeftValue = round(leftCoil[coilIndex] - i[2], 3)
            newLeft.append(newLeftValue)
        else:
            coilIndex2 = actualIndex // 2 - 3
            newRightValue = round(rightCoil[coilIndex2] - i[2], 3)
            newRight.append(newRightValue)

    coilIndex = 10
    coilIndex2 = 10

    for index, i in enumerate(voltageValues[25:40]):
        if coilIndex < len(leftCoil):
            if index % 3 == 0:
                newLeftValue = round(leftCoil[coilIndex] - i[2], 3)
                newLeft.append(newLeftValue)
                coilIndex += 1
        if index >= 2 and (index + 1) % 3 == 0:
            newRightValue = round(rightCoil[coilIndex2] - i[2], 3)
            newRight.append(newRightValue)
            coilIndex2 += 1

    coilIndex = 15
    coilIndex2 = 15

    for index, i in enumerate(voltageValues[40:], start=40):
        if coilIndex < len(leftCoil):
            if (index - 40) % 2 == 0:
                newLeftValue = round(leftCoil[coilIndex] - i[2], 3)
                newLeft.append(newLeftValue)
                coilIndex += 1
        if coilIndex2 < len(rightCoil):
            if (index - 40) % 2 != 0:
                newRightValue = round(rightCoil[coilIndex2] - i[2], 3)
                newRight.append(newRightValue)
                coilIndex2 += 1

    # Pasting the values onto the certificate

    for index, row in enumerate(range(27, 47)):
        sheet.cell(row=row, column=5).value = newFL[index]
        sheet.cell(row=row, column=13).value = newLL[index]
        sheet.cell(row=row, column=9).value = newPF[index]
        sheet.cell(row=row, column=18).value = newPPF[index]
        sheet.cell(row=row, column=10).value = newLeft[index]
        sheet.cell(row=row, column=12).value = newRight[index]

        average = round(
            (
                abs(newFL[index])
                + abs(newLL[index])
                + abs(newPF[index])
                + abs(newLeft[index])
                + abs(newRight[index])
            )
            / 5,
            2,
        )
        averageValues.append(average)
        sheet.cell(row=row, column=23).value = averageValues[index]

    certificate.remove(certificate["6312-3P"])
    certificate.remove(certificate["6312"])
    certificate.remove(certificate["6303"])
    certificate.remove(certificate["Gateway1"])
    certificate.remove(certificate["Gateway2"])
    certificate.save(r"outputs\modifiedCert.xlsx")
    
    openEmptyCert.close()
    certificate.close()


def cert6303(voltage, product, address, regNum, customer, meterTest, xmlFile, csvFile, badgeNum, firmware):
    newCertnum = certificateNumber()


    openEmptyCert = load_workbook(r"excelFiles\emptyCertificate.xlsx")
    sheetCert = openEmptyCert.active

    
    if sheetCert.cell(row=10, column=4).value == "RX-30-430Xytronic":
        bench = "WECO4150"

    elif sheetCert.cell(row=10, column=4).value == "RM-20-102": 
        bench = "WECO2350"

    seriesData, voltageData, voltageData3, seriesData3 = load6312Error(voltage, bench)

    (
        sequenceValue,
        meterID,
        serialNum,
        modelType,
        leftCoil,
        rightCoil,
        PFValue,
        FLValue,
        LLValue,
    ) = raw6303(csvFile)

    macAddress, modelNum, unitSerial, meterConfig, demandInt, pulseType = load6303XML(
        xmlFile
    )

    if str(serialNum[0]) not in unitSerial:
        return print("Serial Numbers don't match")

    newCertnum = certificateNumber()

    certificate = load_workbook(r"excelFiles\emptyCertificate.xlsx")
    sheet = certificate["6303"]

    # Certificate Number
    sheet.cell(row=2, column=4).value = f"{newCertnum}"

    # Owner/Customer, Address and Seal Expiry
    if customer == "Metergy":
        sheet.cell(row=3, column=4).value = "Metergy Solutions Inc."
        sheet.cell(
            row=3, column=16
        ).value = "Suite 601, 8133 Warden Ave., Markham, ON, Canada, L6G 1B3"
        sheet.cell(row=2, column=25).value = "90902"

    else:
        sheet.cell(row=3, column=4).value = customer
        sheet.cell(row=3, column=16).value = address
        sheet.cell(row=2, column=25).value = regNum

    # Verification & Seal Expiry Date

    todayDate = datetime.now()
    currentYear = int(todayDate.year)

    if meterTest == "Verified":
        sheet.cell(row=22, column=4).value = "Verified"
        newYear = currentYear + 10
        sheet.cell(row=2, column=16).value = newYear

    elif meterTest == "Re-verified":
        sheet.cell(row=22, column=4).value = "Re-Verified"
        newYear = currentYear + 8
        sheet.cell(row=2, column=16).value = newYear

    # MAC Address, Serial Number, Model Number, Pulse Weight, Register Type (Display)

    sheet.cell(row=7, column=23).value = macAddress

    sheet.cell(row=5, column=16).value = unitSerial
    sheet.cell(row=5, column=21).value = unitSerial

    for row in range(26, 29):
        sheet.cell(row=row, column=4).value = unitSerial

    sheet.cell(row=7, column=4).value = modelNum

    sheet.cell(row=8, column=9).value = "Display"

    if voltage == "347":
        sheet.cell(row=8, column=4).value = "0.0012"
    elif voltage == "480":
        sheet.cell(row=8, column=4).value = "0.0016"
    elif voltage == "600":
        sheet.cell(row=8, column=4).value = "0.002"
    else:
        sheet.cell(row=8, column=4).value = "0.0004"

    # Output Pulse
    sheet.cell(row=8, column=23).value = "N/A"

    # Element, Current, Voltage
    sheet.cell(row=6, column=21).value = voltage + "V"
    sheet.cell(row=6, column=25).value = "0.80-80mA"

    # Badge Number
    sheet.cell(row=5, column=4).value = badgeNum + "-01"
    sheet.cell(row=5, column=8).value = badgeNum + "-03"

    startNum = 1
    for row in range(26, 29):
        newBadge = f"{badgeNum}-{startNum:02}"
        sheet.cell(row=row, column=3).value = newBadge
        startNum += 1

    # Section E / Test Passing

    sheet.cell(row=22, column=9).value = "Passed"
    sheet.cell(row=22, column=25).value = "Passed"
    sheet.cell(row=23, column=4).value = "Passed"
    sheet.cell(row=23, column=25).value = "Passed"

    # Error Corrections
    sheet.cell(row=17, column=5).value = "Y"
    sheet.cell(row=17, column=6).value = "Y"
    sheet.cell(row=17, column=7).value = "Y"
    sheet.cell(row=17, column=9).value = "Y"
    sheet.cell(row=17, column=19).value = "Y"

    # Config and Element

    if product == "6303-1P":
        sheet.cell(row=7, column=14).value = "1P-03 3W 1PH (2EL)"
        sheet.cell(row=6, column=16).value = 2
    elif product == "6303-2P":
        sheet.cell(row=7, column=14).value = "2P-03 3W Network (2EL)"
        sheet.cell(row=6, column=16).value = 2
    elif product == "6303-3P":
        sheet.cell(row=7, column=14).value = "3P-02 4W Y (3EL)"
        sheet.cell(row=6, column=16).value = 3

    # Firmware Version

    sheet.cell(row=8, column=16).value = firmware

    # Bottom Calculations

    newFL = []
    newLL = []
    newPF = []
    newLeft = []
    newRight = []
    averageValues = []

    if product == "6303-1P" or product == "6303-2P":
        seriesValues = seriesData[voltage][:3]
        voltageValues = voltageData[voltage][:6]
    elif product == "6303-3P":
        seriesValues = seriesData3[voltage][:3]
        voltageValues = voltageData3[voltage][:6]

    for i in seriesValues:
        newValueFL = round(FLValue[seriesValues.index(i)] - i[1], 2)
        newFL.append(newValueFL)

        newValueLL = round(LLValue[seriesValues.index(i)] - i[0], 2)
        newLL.append(newValueLL)

        newValuePF = round(PFValue[seriesValues.index(i)] - i[2], 2)
        newPF.append(newValuePF)

    for index, i in enumerate(voltageValues):
        if index % 2 == 0:
            newLeftValue = round(leftCoil[index // 2] - i[2], 2)
            newLeft.append(newLeftValue)

        else:
            newRightValue = round(rightCoil[index // 2] - i[2], 2)
            newRight.append(newRightValue)

    for index, row in enumerate(range(26, 29)):
        sheet.cell(row=row, column=5).value = newFL[index]
        sheet.cell(row=row, column=13).value = newLL[index]
        sheet.cell(row=row, column=9).value = newPF[index]
        sheet.cell(row=row, column=10).value = newLeft[index]
        sheet.cell(row=row, column=12).value = newRight[index]

        average = round(
            (
                abs(newFL[index])
                + abs(newLL[index])
                + abs(newPF[index])
                + abs(newLeft[index])
                + abs(newRight[index])
            )
            / 5,
            2,
        )
        averageValues.append(average)
        sheet.cell(row=row, column=23).value = averageValues[index]

    certificate.remove(certificate["6312-3P"])
    certificate.remove(certificate["6312"])
    certificate.remove(certificate["6320"])
    certificate.remove(certificate["Gateway1"])
    certificate.remove(certificate["Gateway2"])
    certificate.save(r"outputs\modifiedCert.xlsx")

    openEmptyCert.close()
    certificate.close()


def certGateway(voltage, customer, address, regNum, meterTest, jsonFile, csvFile, badgeNum):
    newCertNum = certificateNumber()

    voltageData, seriesData = loadGateway(voltage)

    macAddress, firmware, modelNum, unitSerial, partModule, serialModule = gatewayJson(jsonFile)

    emptyCert = load_workbook(r"excelFiles\emptyCertificate.xlsx")

    page1 = emptyCert['Gateway1']
    page2 = emptyCert['Gateway2']


    # Certificate Number
    page1.cell(row=2, column=4).value = f"{newCertNum}"
    page2.cell(row=2, column=4).value = f"{newCertNum}"

    # Owner/Customer, Address and Seal Expiry
    if customer == "Metergy":
        page1.cell(row=3, column=4).value = "Metergy Solutions Inc."
        page2.cell(row=3, column=4).value = "Metergy Solutions Inc."
        page1.cell(row=3, column=16).value = "Suite 601, 8133 Warden Ave., Markham, ON, Canada, L6G 1B3"
        page2.cell(row=3, column=16).value = "Suite 601, 8133 Warden Ave., Markham, ON, Canada, L6G 1B3"
        page1.cell(row=2, column=25).value = "90902"
        page2.cell(row=2, column=25).value = "90902"

    else:
        page1.cell(row=3, column=4).value = customer
        page2.cell(row=3, column=4).value = customer
        page1.cell(row=3, column=16).value = address
        page2.cell(row=3, column=16).value = address
        page1.cell(row=2, column=25).value = regNum
        page2.cell(row=2, column=25).value = regNum

    # Verification & Seal Expiry Date

    todayDate = datetime.now()
    currentYear = int(todayDate.year)

    if meterTest == "Verified":
        page1.cell(row=22, column=4).value = "Verified"
        page2.cell(row=22, column=4).value = "Verified"
        newYear = currentYear + 10
        page1.cell(row=2, column=16).value = newYear
        page2.cell(row=2, column=16).value = newYear

    elif meterTest == "Re-verified":
        page1.cell(row=22, column=4).value = "Re-Verified"
        page2.cell(row=22, column=4).value = "Re-Verified"
        newYear = currentYear + 8
        page1.cell(row=2, column=16).value = newYear
        page2.cell(row=2, column=16).value = newYear

     # MAC Address, Serial Number, Model Number, Pulse Weight, Register Type (Display)

    page1.cell(row=7, column=23).value = macAddress
    page2.cell(row=7, column=23).value = macAddress

    page1.cell(row=5, column=16).value = unitSerial
    page2.cell(row=5, column=16).value = unitSerial
    page1.cell(row=5, column=21).value = unitSerial
    page2.cell(row=5, column=21).value = unitSerial

    for row in range(26, 38):
        page1.cell(row=row, column=4).value = unitSerial
        page2.cell(row=row, column=4).value = unitSerial

    page1.cell(row=7, column=4).value = modelNum
    page2.cell(row=7, column=4).value = modelNum

    page1.cell(row=8, column=9).value = "Display"
    page2.cell(row=8, column=9).value = "Display"

    if voltage == "120":
        page1.cell(row=8, column=4).value = "0.0012"
        page2.cell(row=8, column=4).value = "0.0012"
    elif voltage == "480":
        page1.cell(row=8, column=4).value = "0.0016"
        page2.cell(row=8, column=4).value = "0.0016"
    elif voltage == "600":
        page1.cell(row=8, column=4).value = "0.002"
        page2.cell(row=8, column=4).value = "0.002"
    else:
        page1.cell(row=8, column=4).value = "0.0004"
        page2.cell(row=8, column=4).value = "0.0004"

    # Output Pulse
    page1.cell(row=8, column=23).value = "N/A"
    page2.cell(row=8, column=23).value = "N/A"

    # Element, Current, Voltage
    page1.cell(row=6, column=21).value = voltage + "V"
    page2.cell(row=6, column=21).value = voltage + "V"
    page1.cell(row=6, column=25).value = "0.80-80mA"
    page2.cell(row=6, column=25).value = "0.80-80mA"

    # Badge Number
    page1.cell(row=5, column=4).value = badgeNum + "-01"
    page2.cell(row=5, column=4).value = badgeNum + "-13"
    page1.cell(row=5, column=8).value = badgeNum + "-12"
    page2.cell(row=5, column=8).value = badgeNum + "-24"

    startNum = 1
    for row in range(26, 38):
        newBadge = f"{badgeNum}-{startNum:02}"
        page1.cell(row=row, column=3).value = newBadge
        startNum += 1

    for row in range(26, 38):
        newBadge = f"{badgeNum}-{startNum:02}"
        page2.cell(row=row, column=3).value = newBadge
        startNum += 1

    # Firmware 
    page1.cell(row=8, column=16).value = firmware
    page2.cell(row=8, column=16).value = firmware

        # Section E / Test Passing

    page1.cell(row=22, column=9).value = "Passed"
    page1.cell(row=22, column=25).value = "Passed"
    page1.cell(row=23, column=4).value = "Passed"
    page1.cell(row=23, column=25).value = "Passed"

    page2.cell(row=22, column=9).value = "Passed"
    page2.cell(row=22, column=25).value = "Passed"
    page2.cell(row=23, column=4).value = "Passed"
    page2.cell(row=23, column=25).value = "Passed"

    # Error Corrections
    page1.cell(row=17, column=5).value = "Y"
    page1.cell(row=17, column=6).value = "Y"
    page1.cell(row=17, column=7).value = "Y"
    page1.cell(row=17, column=9).value = "Y"
    page1.cell(row=17, column=19).value = "Y"

    page2.cell(row=17, column=5).value = "Y"
    page2.cell(row=17, column=6).value = "Y"
    page2.cell(row=17, column=7).value = "Y"
    page2.cell(row=17, column=9).value = "Y"
    page2.cell(row=17, column=19).value = "Y"


    # Section E Calculations

    emptyCert.remove(emptyCert["6312-3P"])
    emptyCert.remove(emptyCert["6312"])
    emptyCert.remove(emptyCert["6320"])
    emptyCert.remove(emptyCert["6303"])
    emptyCert.save(r"outputs\modifiedCert.xlsx")


    emptyCert.close()


