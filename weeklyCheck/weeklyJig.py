from openpyxl import load_workbook
import win32api
from datetime import date, datetime
from openpyxl.styles import PatternFill
import os
 
def exportDataJig(dataFile):
    # This is the function that will grab the errors and store them. The user submits this file
    data = load_workbook(dataFile)
    dataSheet = data.active

    dataA = []
    dataB = []
    dataC = []
    Amps = []
    factorValue = []
    maxRows = 0
    
    for rows in range(6, dataSheet.max_row + 1):
        if dataSheet.cell(row=rows, column=1).value is None or dataSheet.cell(row=rows, column=1).value == "":
            maxRows = rows
            break

    for row in range(1, maxRows):
        cell2 = dataSheet.cell(row=row, column=2).value
        cell4 = dataSheet.cell(row=row, column=4).value
        cell8 = dataSheet.cell(row=row + 5, column=8).value
        cell3 = dataSheet.cell(row=row + 5, column=3).value

        if cell2 == "A":
            dataA.append(cell4)
           
        elif cell2 == "B":
            dataB.append(cell4)
        
        elif cell2 == "C":
            dataC.append(cell4)

        if dataSheet.cell(row=row + 5, column=4).value:
            Amps.append(cell8)

        else:
            Amps.append(None)

        if cell3 == "CR":
            factorValue.append(None)

        else:
            factorValue.append(cell3)

    data.close()
    return dataA, dataB, dataC, Amps, factorValue


def exportErrors(bench, voltage, prodType):
    
    # Type is either 6320, 6312-XP, 6312-3P, GT-XP-01

    workbook = load_workbook(r"weeklyCheck\excelFile\TestJigCert.xlsx")
    
    if bench == "WECO4150":
        sheet = workbook['Errors WECO4150']

    elif bench == "WECO2350":
        sheet = workbook['Errors WECO2350']

    if prodType == "6320":
        voltage = "120"
    
    econConsole = []
    econRadian = []

    if bench == "WECO4150":
        if prodType == '6312-XP' or prodType == '6312-3P':
            voltageMap = {
                "120": 120,
                "240": 240,
                "277": 277,
                "347": 347,
                "480": 480,
                "600": 600
            }
        
            # TRIACTA-6312-JIG-XP-02
            voltageValue = voltageMap.get(voltage)

            if prodType == '6312-XP':
                if voltageValue is not None:
                    for row in range(8, 589):
                        econConsoleCell = sheet.cell(row=row, column=6).value
                        econRadianCell = sheet.cell(row=row, column=7).value
                        voltageCell = sheet.cell(row=row, column=1).value

                        if voltageCell == voltageValue:
                            econConsole.append(econConsoleCell)
                            econRadian.append(econRadianCell)
        
            elif prodType == '6312-3P':
                if voltageValue is not None:
                    for row in range(8, 584):
                        econConsoleCell = sheet.cell(row=row, column=15).value
                        econRadianCell = sheet.cell(row=row, column=16).value
                        voltageCell = sheet.cell(row=row, column=10).value

                        if voltageCell == voltageValue:
                            econConsole.append(econConsoleCell)
                            econRadian.append(econRadianCell)
            
        if prodType == '6320':
            for row in range(8, 208):
                econConsole.append(sheet.cell(row=row,column=24).value)
                econRadian.append(sheet.cell(row=row,column=25).value)
    
    elif bench == "WECO2350":
        if prodType == '6312-XP' or prodType == '6312-3P' or prodType == "GT-XP-01":
            voltageMap = {
                "120": 120,
                "347": 347,
                "600": 600
            }
        
            voltageValue = voltageMap.get(voltage)

            if prodType == '6312-XP':
                if voltageValue is not None:
                    for row in range(8, 200):
                        econConsoleCell = sheet.cell(row=row, column=15).value
                        econRadianCell = sheet.cell(row=row, column=16).value
                        voltageCell = sheet.cell(row=row, column=10).value

                        if voltageCell == voltageValue:
                            econConsole.append(econConsoleCell)
                            econRadian.append(econRadianCell)
        
            elif prodType == '6312-3P':
                if voltageValue is not None:
                    for row in range(8, 200):
                        econConsoleCell = sheet.cell(row=row, column=24).value
                        econRadianCell = sheet.cell(row=row, column=25).value
                        voltageCell = sheet.cell(row=row, column=19).value

                        if voltageCell == voltageValue:
                            econConsole.append(econConsoleCell)
                            econRadian.append(econRadianCell)

            elif prodType == "GT-XP-01":
                if voltageValue is not None:
                    for row in range(8, 586):
                        econConsoleCell = sheet.cell(row=row, column=33).value
                        econRadianCell = sheet.cell(row=row, column=34).value
                        voltageCell = sheet.cell(row=row, column=28).value

                        if voltageCell == voltageValue:
                            econConsole.append(econConsoleCell)
                            econRadian.append(econRadianCell)
            
        if prodType == '6320':
            for row in range(8, 208):
                econConsole.append(sheet.cell(row=row,column=6).value)
                econRadian.append(sheet.cell(row=row,column=7).value)

    workbook.close()
    return econConsole, econRadian


def generateTestJigCert(roomTemp, dataFile, bench, voltage, prodType):
    emptyCert = load_workbook(r"weeklyCheck\excelFile\TestJigCert.xlsx")
    sheet = emptyCert['Cert']

    # Usernames, changing info

    verifier = win32api.GetUserName()

    # Add more to verifier if needed
    # Verifier = Metergy Username
    if verifier == "alachhman":
        sheet.cell(row=10, column=3).value = "Ashraff Lachman"
        sheet.cell(row=10, column=9).value = "Ashraff Lachman"
    elif verifier == "rchampaneri":
        sheet.cell(row=10, column=3).value = "Raj Champaneri"
        sheet.cell(row=10, column=9).value = "Raj Champaneri"
    elif verifier == "adifebo":
        sheet.cell(row=10, column=3).value = "Adam Di Febo"
        sheet.cell(row=10, column=9).value = "Adam Di Febo"
    elif verifier == "kneild2":
        sheet.cell(row=10, column=3).value = "Kaden Neild"
        sheet.cell(row=10, column=9).value = "Kaden Neild"

    sheet.cell(row=11, column=3).value = roomTemp + "C"

    dateValue = date.today().strftime("%d-%m-%Y")
    sheet.cell(row=11, column=6).value = dateValue

    if bench == "WECO2350":
        if prodType == "6320":
            sheet.cell(row=2, column=4).value = "TRIACTA-6320-JIG-XP-01"
        elif prodType == "6312-XP":
            sheet.cell(row=2, column=4).value = "TRIACTA-6312-JIG-XP-02"
        elif prodType == "6312-3P":
            sheet.cell(row=2, column=4).value = "TRIACTA-6312-JIG-3P-02"
        elif prodType == "GT-XP-01":
            sheet.cell(row=2, column=4).value = "TRIACTA-GATEWAY-GT-XP-01"				
        
        # Test Jig Data 
        sheet.cell(row=6, column=4).value = "206351"
        sheet.cell(row=5, column=6).value = "6500"
        sheet.cell(row=6, column=6).value = "206352"
        sheet.cell(row=5, column=8).value = "MD-CC-02"
        sheet.cell(row=6, column=8).value = "206353"
        sheet.cell(row=5, column=10).value = "2021-12-20"
        sheet.cell(row=6, column=10).value = "2021-12-20"

    elif bench == "WECO4150":
        if prodType == "6320":
            sheet.cell(row=2, column=4).value = "TRIACTA-6320-JIG-3P-01"
        elif prodType == "6312-XP":
            sheet.cell(row=2, column=4).value = "TRIACTA-6312-JIG-XP-01"
        elif prodType == "6312-3P":
            sheet.cell(row=2, column=4).value = "TRIACTA-6312-JIG-3P-01"

        # Test Jig Data 
        sheet.cell(row=6, column=4).value = "710194"
        sheet.cell(row=5, column=6).value = "8777"
        sheet.cell(row=6, column=6).value = "710194"
        sheet.cell(row=5, column=8).value = "MD-CC-03"
        sheet.cell(row=6, column=8).value = "710194"
        sheet.cell(row=5, column=10).value = "2022-05-27"
        sheet.cell(row=6, column=10).value = "2022-05-27"


    # Calculations

    dataA, dataB, dataC, Amps, factorValue = exportDataJig(dataFile)
    econConsole, econRadian = exportErrors(bench, voltage, prodType)

    row = 19
    numberA = 1
    numberB = 1
    numberC = 1

    numberACol = PatternFill(start_color="ffb4c6e7", end_color="ffb4c6e7", fill_type="solid")
    numberBCol = PatternFill(start_color="fff4b084", end_color="fff4b084", fill_type="solid")
    numberCCol = PatternFill(start_color="ffa9d08e", end_color="ffa9d08e", fill_type="solid")

    for value in dataA:
        if value is not None:
            sheet.cell(row=row, column=8).value = 100 - value
            sheet.cell(row=row, column=5).value = f"P{numberA}-1"
            sheet.cell(row=row, column=5).fill = numberACol
            row += 1
        else:
            row += 1
            numberA += 1 

    for value in dataB:
        if value is not None:
            sheet.cell(row=row, column=8).value = 100 - value
            sheet.cell(row=row, column=5).value = f"P{numberB}-2"
            sheet.cell(row=row, column=5).fill = numberBCol
            row += 1           
        else:
            row += 1
            numberB += 1 

    for value in dataC:
        if value is not None:
            sheet.cell(row=row, column=8).value = 100 - value
            sheet.cell(row=row, column=5).value = f"P{numberC}-3"
            sheet.cell(row=row, column=5).fill = numberCCol
            row += 1
        else:
            row += 1 
            numberC += 1

    row = 19

    for value in Amps:
        if value is not None:
            sheet.cell(row=row, column=2).value = value
            row += 1
        else:
            row += 1 

    row = 19

    for value in factorValue:
        if sheet.cell(row=row, column=8).value is not None:
            if value is not None:
                sheet.cell(row=row, column=3).value = value
                row += 1

        else:
            row += 1 

    row = 19

    i = 0
    while i < len(econConsole):
        if sheet.cell(row=row, column=8).value is not None:
            sheet.cell(row=row, column=6).value = econConsole[i]
            sheet.cell(row=row, column=1).value = voltage
            sheet.cell(row=row, column=4).value = "Delivered"
            i += 1       
        row += 1

    row = 19

    i = 0 
    while i < len(econRadian):
        if sheet.cell(row=row, column=8).value is not None:
            sheet.cell(row=row, column=7).value = econRadian[i]
            i += 1    
        row += 1

    emptyCert.remove(emptyCert['Errors WECO4150'])
    emptyCert.remove(emptyCert['Errors WECO2350'])
    emptyCert.save(r"weeklyCheck\excelFile\Modified.xlsx")
    emptyCert.close()


def exportJigCert(meterPoint):
    modifiedCert = load_workbook(r"weeklyCheck\excelFile\Modified.xlsx")
    sheet = modifiedCert.active

    product = sheet.cell(row=2, column=4).value
    testboardNum = sheet.cell(row=5, column=8).value
    dateValue = sheet.cell(row=11, column=6).value

    try:
        if isinstance(dateValue, str):
            date_obj = datetime.strptime(dateValue, "%d-%m-%Y")
        elif isinstance(dateValue, datetime):
            date_obj = dateValue
        else:
            raise ValueError(f"Date value '{dateValue}' is not in expected format")

        dateValue = date_obj.strftime("%Y-%m-%d")
    
    except ValueError as e:
        print(f"Error parsing date: {e}")
        return  

    fileName = f"{dateValue}_{testboardNum}_{product}_{meterPoint}.xlsx"

    savePath = os.path.join("weeklyCheck", "completed", fileName)

    modifiedCert.close()
    os.remove(r"weeklyCheck\excelFile\Modified.xlsx")

    modifiedCert.save(savePath)

