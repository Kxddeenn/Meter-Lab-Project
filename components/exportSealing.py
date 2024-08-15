from openpyxl import load_workbook
from datetime import date
import win32api
import os


def exportSealing():
    """
    Export data from 'modifiedCert.xlsx' to 'SealingLog.xlsx' and save as individual completed certificate.
    """
    # Load 'modifiedCert.xlsx' to read data from
    modifiedCert = load_workbook(r"outputs\modifiedCert.xlsx")
    sheet = modifiedCert.active

    # Load 'SealingLog.xlsx' to append new sealing entry
    sealingLog = load_workbook(r"excelFiles\SealingLog.xlsx")
    sealing = sealingLog["2024"]

    # Retrieve values from 'modifiedCert.xlsx'
    badgeNum = sheet.cell(row=5, column=4).value
    certNum = sheet.cell(row=2, column=4).value
    config = sheet.cell(row=7, column=4).value

    noaValue = None

    badgeNum = badgeNum.split("-")[0]

    # Determine noaValue based on config
    if "20" in config:
        noaValue = "AE-1434"
    elif "12" in config:
        noaValue = "AE-1665"

    serialNumber = sheet.cell(row=5, column=16).value
    macNumber = sheet.cell(row=7, column=23).value

    serialNumber = serialNumber[1:]

    # Determine typeLetter based on demandTest
    demandTest = sheet.cell(row=23, column=16).value
    typeLetter = "D" if demandTest == "Passed" else "E"

    # Get today's date
    dateValue = date.today().strftime("%B %d, %Y")

    # Get current user's username
    verifier = win32api.GetUserName()

    firmwareNumber = sheet.cell(row=8, column=16).value
    verification = sheet.cell(row=22, column=4).value
    testConsole = sheet.cell(row=11, column=4).value
    contractor = sheet.cell(row=3, column=4).value
    regNum = sheet.cell(row=2, column=25).value
    customer = sheet.cell(row=3, column=4).value


    # Find next available row in 'SealingLog.xlsx'
    maxRow = sealing.max_row
    while maxRow > 0 and sealing.cell(row=maxRow, column=3).value is None:
        maxRow -= 1
    nextRow = maxRow + 1

    if "Metergy" in customer:
        sealing.cell(row=nextRow, column=1).value = "Metergy Requests"
    else:
        sealing.cell(row=nextRow, column=1).value = "Please input manually"

    # Write data to 'SealingLog.xlsx'
    sealing.cell(row=nextRow, column=2).value = badgeNum
    sealing.cell(row=nextRow, column=3).value = certNum
    sealing.cell(row=nextRow, column=4).value = config
    sealing.cell(row=nextRow, column=5).value = noaValue
    sealing.cell(row=nextRow, column=6).value = serialNumber
    sealing.cell(row=nextRow, column=7).value = macNumber
    sealing.cell(row=nextRow, column=8).value = typeLetter
    sealing.cell(row=nextRow, column=9).value = dateValue
    sealing.cell(row=nextRow, column=10).value = verifier
    sealing.cell(row=nextRow, column=11).value = firmwareNumber
    sealing.cell(row=nextRow, column=12).value = verification
    sealing.cell(row=nextRow, column=14).value = testConsole
    sealing.cell(row=nextRow, column=15).value = contractor
    sealing.cell(row=nextRow, column=16).value = regNum
    sealing.cell(row=nextRow, column=17).value = "Passed"

    # Save changes to 'SealingLog.xlsx'
    sealingLog.save(r"excelFiles\SealingLog.xlsx")

    # Save 'modifiedCert.xlsx' as completed certificate in 'completed' folder
    modifiedCert.save(f"completed/{certNum}.xlsx")

    modifiedCert.close()
    sealingLog.close()

    # Optionally, remove 'modifiedCert.xlsx' (uncomment to enable)
    os.remove(r"outputs\modifiedCert.xlsx")


