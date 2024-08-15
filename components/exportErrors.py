from openpyxl import load_workbook

def load6312Error(voltageSelection, bench):
    """
    Loads error data for the 6312 product series based on voltage selection and test bench type.

    Args:
        voltageSelection (str): Voltage type ('120', '240', '277', '347', '480', '600').
        bench (str): Type of test bench ('WECO4150' or 'WECO2350').

    Returns:
        tuple: Tuple containing dictionaries for series and voltage data for 6312-2P and 6312-3P sheets.
    """

    if bench == 'WECO4150':
        errorWorkbook = load_workbook(r"excelFiles\WECO4150_CertErrors.xlsx", data_only=True)

    elif bench == 'WECO2350':
        errorWorkbook = load_workbook(r"excelFiles\WECO2350_CertErrors.xlsx", data_only=True)

    # Load 6312-2P sheet
    product6312Sheet = errorWorkbook['6312-2P']


    # Define columns for voltage and series data for 6312-2P sheet
    voltageColumns = {
        '120': {'LL': 'D', 'FL': 'E', 'PF': 'F'},
        '240': {'LL': 'K', 'FL': 'L', 'PF': 'M'},
        '277': {'LL': 'R', 'FL': 'S', 'PF': 'T'},
        '347': {'LL': 'Y', 'FL': 'Z', 'PF': 'AA'},
        '480': {'LL': 'AF', 'FL': 'AG', 'PF': 'AH'},
        '600': {'LL': 'AM', 'FL': 'AN', 'PF': 'AO'}
    }

    seriesColumns = {
        '120': {'LL': 'G', 'FL': 'H', 'PF': 'I'},
        '240': {'LL': 'N', 'FL': 'O', 'PF': 'P'},
        '277': {'LL': 'U', 'FL': 'V', 'PF': 'W'},
        '347': {'LL': 'AB', 'FL': 'AC', 'PF': 'AD'},
        '480': {'LL': 'AI', 'FL': 'AJ', 'PF': 'AK'},
        '600': {'LL': 'AP', 'FL': 'AQ', 'PF': 'AR'} 
    }

    voltageData = {}
    seriesData = {}

    # Process data for 6312-2P sheet
    if voltageSelection is not None and voltageSelection not in voltageColumns:
        raise ValueError(f"Invalid voltage type: {voltageSelection}")

    for voltageType, columns in voltageColumns.items():
        
        if voltageSelection is not None and voltageType != voltageSelection:
            continue

        ll_column = columns['LL']
        fl_column = columns['FL']
        pf_column = columns['PF']

        ll_values = []
        fl_values = []
        pf_values = []

        for row in range(8, 32):
            ll_value = product6312Sheet[f"{ll_column}{row}"].value
            fl_value = product6312Sheet[f"{fl_column}{row}"].value
            pf_value = product6312Sheet[f"{pf_column}{row}"].value

            ll_values.append(ll_value)
            fl_values.append(fl_value)
            pf_values.append(pf_value)

        combined_data = list(zip(ll_values, fl_values, pf_values))
        voltageData[voltageType] = combined_data

        if voltageSelection is not None:
            break

    


    for seriesType, columns in seriesColumns.items():

        if voltageSelection is not None and seriesType != voltageSelection:
            continue


        ll_column2 = columns['LL']
        fl_column2 = columns['FL']
        pf_column2 = columns['PF']

        ll_series = []
        fl_series = []
        pf_series = []

        for row in range(8, 31):
            ll_seriesValue = product6312Sheet[f"{ll_column2}{row}"].value
            fl_seriesValue = product6312Sheet[f"{fl_column2}{row}"].value
            pf_seriesValue = product6312Sheet[f"{pf_column2}{row}"].value

            if isinstance(ll_seriesValue, (int, float)) and isinstance(fl_seriesValue, (int, float)) and isinstance(pf_seriesValue, (int, float)):
                ll_series.append(round(ll_seriesValue, 3))
                fl_series.append(round(fl_seriesValue, 3))
                pf_series.append(round(pf_seriesValue, 3))

            else:     
                ll_series.append('')
                fl_series.append('')
                pf_series.append('')

        combinedSeries = list(zip(ll_series, fl_series, pf_series))
        combinedSeries = [tpl for tpl in combinedSeries if any(val != '' for val in tpl)]
        seriesData[seriesType] = combinedSeries

        if voltageSelection is not None:
            break

    # Load 6312-3P sheet
    product6312Sheet = errorWorkbook['6312-3P']

    # Define columns for voltage and series data for 6312-3P sheet
    voltageColumns3 = {
        '120': {'LL': 'D', 'FL': 'E', 'PF': 'F'},
        '240': {'LL': 'K', 'FL': 'L', 'PF': 'M'},
        '277': {'LL': 'R', 'FL': 'S', 'PF': 'T'},
        '347': {'LL': 'Y', 'FL': 'Z', 'PF': 'AA'},
        '480': {'LL': 'AF', 'FL': 'AG', 'PF': 'AH'},
        '600': {'LL': 'AM', 'FL': 'AN', 'PF': 'AO'}
    }

    seriesColumns3 = {
        '120': {'LL': 'G', 'FL': 'H', 'PF': 'I'},
        '240': {'LL': 'N', 'FL': 'O', 'PF': 'P'},
        '277': {'LL': 'U', 'FL': 'V', 'PF': 'W'},
        '347': {'LL': 'AB', 'FL': 'AC', 'PF': 'AD'},
        '480': {'LL': 'AI', 'FL': 'AJ', 'PF': 'AK'},
        '600': {'LL': 'AP', 'FL': 'AQ', 'PF': 'AR'} 
    }

    seriesData3 = {}
    voltageData3 = {}

    # Process data for 6312-3P sheet
    if voltageSelection is not None and voltageSelection not in voltageColumns3:
        raise ValueError(f"Invalid voltage type: {voltageSelection}")

    for voltageType, columns in voltageColumns.items():
        
        if voltageSelection is not None and voltageType != voltageSelection:
            continue

        ll_column = columns['LL']
        fl_column = columns['FL']
        pf_column = columns['PF']

        ll_values = []
        fl_values = []
        pf_values = []

        for row in range(8, 32):
            ll_value = product6312Sheet[f"{ll_column}{row}"].value
            fl_value = product6312Sheet[f"{fl_column}{row}"].value
            pf_value = product6312Sheet[f"{pf_column}{row}"].value

            ll_values.append(ll_value)
            fl_values.append(fl_value)
            pf_values.append(pf_value)

        combined_data = list(zip(ll_values, fl_values, pf_values))
        voltageData3[voltageType] = combined_data

        if voltageSelection is not None:
            break

    for seriesType, columns in seriesColumns3.items():

        if voltageSelection is not None and seriesType != voltageSelection:
            continue


        ll_column2 = columns['LL']
        fl_column2 = columns['FL']
        pf_column2 = columns['PF']

        ll_series = []
        fl_series = []
        pf_series = []

        for row in range(8, 31):
            ll_seriesValue = product6312Sheet[f"{ll_column2}{row}"].value
            fl_seriesValue = product6312Sheet[f"{fl_column2}{row}"].value
            pf_seriesValue = product6312Sheet[f"{pf_column2}{row}"].value

            if isinstance(ll_seriesValue, (int, float)) and isinstance(fl_seriesValue, (int, float)) and isinstance(pf_seriesValue, (int, float)):
                ll_series.append(round(ll_seriesValue, 3))
                fl_series.append(round(fl_seriesValue, 3))
                pf_series.append(round(pf_seriesValue, 3))

            else:     
                ll_series.append('')
                fl_series.append('')
                pf_series.append('')

        combinedSeries = list(zip(ll_series, fl_series, pf_series))
        combinedSeries = [tpl for tpl in combinedSeries if any(val != '' for val in tpl)]
        seriesData3[seriesType] = combinedSeries

        if voltageSelection is not None:
            break

    errorWorkbook.close()

    return seriesData, voltageData, voltageData3, seriesData3

def load6312GatewayErr(voltageSelection, product):
    errorWorkbook = load_workbook(r"excelFiles\WECO2350_CertErrors.xlsx", data_only=True)
    product6312Sheet = errorWorkbook['Gateway']

    voltageColumns = {
        '120': {'LL': 'D', 'FL': 'E', 'PF': 'F'},
        '347': {'LL': 'L', 'FL': 'M', 'PF': 'N'},
        '600': {'LL': 'T', 'FL': 'U', 'PF': 'V'}
    }

    seriesColumns = {
        '120': {'LL': 'H', 'FL': 'I', 'PF': 'J'},
        '347': {'LL': 'P', 'FL': 'Q', 'PF': 'R'},
        '600': {'LL': 'X', 'FL': 'Y', 'PF': 'Z'} 
    }

    voltageData = {}
    seriesData = {}

    if voltageSelection is not None and voltageSelection not in voltageColumns:
        raise ValueError(f"Invalid voltage type: {voltageSelection}")

    for voltageType, columns in voltageColumns.items():
        
        if voltageSelection is not None and voltageType != voltageSelection:
            continue

        ll_column = columns['LL']
        fl_column = columns['FL']
        pf_column = columns['PF']

        ll_values = []
        fl_values = []
        pf_values = []

        if product == '6312 (1-12)':
            for row in range(6, 30):
                ll_value = product6312Sheet[f"{ll_column}{row}"].value
                fl_value = product6312Sheet[f"{fl_column}{row}"].value
                pf_value = product6312Sheet[f"{pf_column}{row}"].value

                ll_values.append(ll_value)
                fl_values.append(fl_value)
                pf_values.append(pf_value)

            combined_data = list(zip(ll_values, fl_values, pf_values))
            voltageData[voltageType] = combined_data

            if voltageSelection is not None:
                break
        elif product == '6312 (13-24)':
            for row in range(30, 54):
                ll_value = product6312Sheet[f"{ll_column}{row}"].value
                fl_value = product6312Sheet[f"{fl_column}{row}"].value
                pf_value = product6312Sheet[f"{pf_column}{row}"].value

                ll_values.append(ll_value)
                fl_values.append(fl_value)
                pf_values.append(pf_value)

            combined_data = list(zip(ll_values, fl_values, pf_values))
            voltageData[voltageType] = combined_data

            if voltageSelection is not None:
                break

    for seriesType, columns in seriesColumns.items():

        if voltageSelection is not None and seriesType != voltageSelection:
            continue


        ll_column2 = columns['LL']
        fl_column2 = columns['FL']
        pf_column2 = columns['PF']

        ll_series = []
        fl_series = []
        pf_series = []

        if product == '6312 (1-12)':
            for row in range(6, 30):
                ll_seriesValue = product6312Sheet[f"{ll_column2}{row}"].value
                fl_seriesValue = product6312Sheet[f"{fl_column2}{row}"].value
                pf_seriesValue = product6312Sheet[f"{pf_column2}{row}"].value

                if isinstance(ll_seriesValue, (int, float)) and isinstance(fl_seriesValue, (int, float)) and isinstance(pf_seriesValue, (int, float)):
                    ll_series.append(round(ll_seriesValue, 3))
                    fl_series.append(round(fl_seriesValue, 3))
                    pf_series.append(round(pf_seriesValue, 3))

                else:     
                    ll_series.append('')
                    fl_series.append('')
                    pf_series.append('')

        elif product == '6312 (13-24)':
            for row in range(30, 54):
                ll_seriesValue = product6312Sheet[f"{ll_column2}{row}"].value
                fl_seriesValue = product6312Sheet[f"{fl_column2}{row}"].value
                pf_seriesValue = product6312Sheet[f"{pf_column2}{row}"].value

                if isinstance(ll_seriesValue, (int, float)) and isinstance(fl_seriesValue, (int, float)) and isinstance(pf_seriesValue, (int, float)):
                    ll_series.append(round(ll_seriesValue, 3))
                    fl_series.append(round(fl_seriesValue, 3))
                    pf_series.append(round(pf_seriesValue, 3))

                else:     
                    ll_series.append('')
                    fl_series.append('')
                    pf_series.append('')


        combinedSeries = list(zip(ll_series, fl_series, pf_series))
        combinedSeries = [tpl for tpl in combinedSeries if any(val != '' for val in tpl)]
        seriesData[seriesType] = combinedSeries

        if voltageSelection is not None:
            break
    
    errorWorkbook.close()

    return seriesData, voltageData

def load6320Error(voltageSelection, seriesSelection, bench):
    """
    Loads error data for the 6320 product series based on voltage and series selection and test bench type.

    Args:
        voltageSelection (str): Voltage type ('120').
        seriesSelection (str): Series type ('6320-3P', '6320-2P', '6320-1P').
        bench (str): Type of test bench ('WECO4150' or 'WECO2350').

    Returns:
        tuple: Tuple containing dictionaries for series and voltage data for the 6320 sheet.
    """
    
    if bench == 'WECO4150':
        errorWorkbook = load_workbook(r"excelFiles\WECO4150_CertErrors.xlsx", data_only=True)

    elif bench == 'WECO2350':
        errorWorkbook = load_workbook(r"excelFiles\WECO2350_CertErrors.xlsx", data_only=True)
    product6312Sheet = errorWorkbook['6320']

    # Define columns for voltage and series data for 6320 sheet
    voltageColumns = {
        '120': {'LL': 'C', 'FL': 'D', 'PF': 'E', 'PPF': 'F'}
    }

    seriesColumns = {
        '6320-3P': {'LL': 'G', 'FL': 'H', 'PF': 'I', 'PPF': 'J'},
        '6320-2P': {'LL': 'K', 'FL': 'L', 'PF': 'M', 'PPF': 'N'},
        '6320-1P': {'LL': 'K', 'FL': 'L', 'PF': 'M', 'PPF': 'N'}
    }

    voltageData = {}
    seriesData = {}

    # Process voltage data for 6320 sheet
    for voltageType, columns in voltageColumns.items():
        
        if voltageSelection is not None and voltageType != voltageSelection:
            continue

        ll_column = columns['LL']
        fl_column = columns['FL']
        pf_column = columns['PF']
        pf2_column = columns['PPF']

        ll_values = []
        fl_values = []
        pf_values = []
        pf2_values = []

        for row in range(3, 53):
            ll_value = product6312Sheet[f"{ll_column}{row}"].value
            fl_value = product6312Sheet[f"{fl_column}{row}"].value
            pf_value = product6312Sheet[f"{pf_column}{row}"].value
            pf2_value = product6312Sheet[f"{pf2_column}{row}"].value

            ll_values.append(ll_value)
            fl_values.append(fl_value)
            pf_values.append(pf_value)
            pf2_values.append(pf2_value)

        combined_data = list(zip(ll_values, fl_values, pf_values, pf2_values))
        voltageData[voltageType] = combined_data

        if voltageSelection is not None:
            break

    # Process series data for 6320 sheet
    if seriesSelection is not None and seriesSelection in seriesColumns:
        columns = seriesColumns[seriesSelection]
        ll_column = columns['LL']
        fl_column = columns['FL']
        pf_column = columns['PF']
        pf2_column = columns['PPF']

        ll_series = []
        fl_series = []
        pf_series = []
        pf2_series = []

        for row in range(2, 53):
            ll_seriesValue = product6312Sheet[f"{ll_column}{row}"].value
            fl_seriesValue = product6312Sheet[f"{fl_column}{row}"].value
            pf_seriesValue = product6312Sheet[f"{pf_column}{row}"].value
            pf2_seriesValue = product6312Sheet[f"{pf2_column}{row}"].value

            if isinstance(ll_seriesValue, (int, float)) and isinstance(fl_seriesValue, (int, float)) and isinstance(pf_seriesValue, (int, float)) and isinstance(pf2_seriesValue, (int, float)):
                ll_series.append(round(ll_seriesValue, 3))
                fl_series.append(round(fl_seriesValue, 3))
                pf_series.append(round(pf_seriesValue, 3))
                pf2_series.append(round(pf2_seriesValue, 3))
            else:     
                ll_series.append('')
                fl_series.append('')
                pf_series.append('')
                pf2_series.append('')

        combinedSeries = list(zip(ll_series, fl_series, pf_series, pf2_series))
        combinedSeries = [tpl for tpl in combinedSeries if any(val != '' for val in tpl)]
        seriesData[seriesSelection] = combinedSeries

    errorWorkbook.close()

    return seriesData, voltageData
    
def loadGateway(voltageSelection):
    """
    Loads error data for the new Gateway product series based on voltage selection. Only on WECO2350 currently.

    Args:
        voltageSelection (str): Voltage type ('120', '347', '600').

    Returns:
        tuple: Tuple containing dictionaries for series and voltage data for the Gateway sheet.
    """

    errorWorkbook = load_workbook(r"excelFiles\WECO2350_CertErrors.xlsx", data_only= True)
    gatewaysheet = errorWorkbook['Gateway']

        
    voltageColumns = {
        '120': {'LL': 'D', 'FL': 'E', 'PF': 'F'},
        '347': {'LL': 'L', 'FL': 'M', 'PF': 'N'},
        '600': {'LL': 'T', 'FL': 'U', 'PF': 'V'}
    }

    seriesColumns = {
        '120': {'LL': 'H', 'FL': 'I', 'PF': 'J'},
        '347': {'LL': 'P', 'FL': 'Q', 'PF': 'R'},
        '600': {'LL': 'X', 'FL': 'Y', 'PF': 'Z'} 
    }

    voltageData = {}
    seriesData = {}

    if voltageSelection is not None and voltageSelection not in voltageColumns:
        raise ValueError(f"Invalid voltage type: {voltageSelection}")

    for voltageType, columns in voltageColumns.items():
        
        if voltageSelection is not None and voltageType != voltageSelection:
            continue

        ll_column = columns['LL']
        fl_column = columns['FL']
        pf_column = columns['PF']

        ll_values = []
        fl_values = []
        pf_values = []

        for row in range(6, 54):
            ll_value = gatewaysheet[f"{ll_column}{row}"].value
            fl_value = gatewaysheet[f"{fl_column}{row}"].value
            pf_value = gatewaysheet[f"{pf_column}{row}"].value

            ll_values.append(ll_value)
            fl_values.append(fl_value)
            pf_values.append(pf_value)

        combined_data = list(zip(ll_values, fl_values, pf_values))
        voltageData[voltageType] = combined_data

        if voltageSelection is not None:
            break

        


    for seriesType, columns in seriesColumns.items():

        if voltageSelection is not None and seriesType != voltageSelection:
            continue


        ll_column2 = columns['LL']
        fl_column2 = columns['FL']
        pf_column2 = columns['PF']

        ll_series = []
        fl_series = []
        pf_series = []

        for row in range(6, 54):
            ll_seriesValue = gatewaysheet[f"{ll_column2}{row}"].value 
            fl_seriesValue = gatewaysheet[f"{fl_column2}{row}"].value
            pf_seriesValue = gatewaysheet[f"{pf_column2}{row}"].value


            if isinstance(ll_seriesValue, (int, float)) and isinstance(fl_seriesValue, (int, float)) and isinstance(pf_seriesValue, (int, float)):
                ll_series.append(round(ll_seriesValue, 3))
                fl_series.append(round(fl_seriesValue, 3))
                pf_series.append(round(pf_seriesValue, 3))

            else:     
                ll_series.append('')
                fl_series.append('')
                pf_series.append('')

        combinedSeries = list(zip(ll_series, fl_series, pf_series))
        combinedSeries = [tpl for tpl in combinedSeries if any(val != '' for val in tpl)]
        seriesData[seriesType] = combinedSeries

        if voltageSelection is not None:
            break

        
    errorWorkbook.close()
    
    return voltageData, seriesData


