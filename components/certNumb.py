from openpyxl import load_workbook

def certificateNumber():
    """
    Generates a new certificate number based on the last entry in an Excel sheet.
    Assumes the Excel file 'SealingLog.xlsx' exists and has valid data.

    Returns:
        str: New certificate number generated based on existing entries.
    """
    sealingLog = load_workbook(r"excelFiles\SealingLog.xlsx")
    sealingSheet = sealingLog['2024']

    maxRow = sealingSheet.max_row
    
    # Find the last non-empty cell in column C
    while maxRow > 0 and sealingSheet.cell(row=maxRow, column=3).value is None:
        maxRow -= 1

    # Get the last certificate number
    certNum = sealingSheet.cell(row=maxRow, column=3).value

    # Split into text and numeric parts
    text, number = certNum.rsplit('-', 1)

    # Increment the numeric part
    newNumber = int(number) + 1

    # Ensure the new number retains the same width as the original
    newNumberStr = str(newNumber).zfill(len(number))

    # Construct the new certificate number
    newCertNum = f"{text}-{newNumberStr}"

    sealingLog.close()

    return newCertNum


