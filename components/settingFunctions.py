from openpyxl import load_workbook
import os
import webbrowser
from dateutil import parser

# ALL UI COMPONENTS

def saveJigDate(date):
    """
    Saves a given date to specific cells in different sheets of the empty certificate.

    Args:
        date (str): The date to be saved in format 'YYYY-MM-DD'.

    Returns:
        None
    """   
    openEmpty = load_workbook(r"excelFiles\emptyCertificate.xlsx")

    sheet6312 = openEmpty['6312']
    sheet6320 = openEmpty['6320']
    sheet6303 = openEmpty['6303']
    sheet36312 = openEmpty['6312-3P']
    Gateway1 = openEmpty['Gateway1']
    Gateway2 = openEmpty['Gateway2']

    try: 
        date = parser.parse(date)
        date = date.strftime("%d-%b-%Y")
    
    except ValueError as e:
        print("Couldn't parse the date properly")

    sheet6312.cell(row=12, column=23).value = date
    sheet6320.cell(row=12, column=23).value = date
    sheet6303.cell(row=12, column=23).value = date
    sheet36312.cell(row=12, column=23).value = date
    Gateway1.cell(row=12, column=23).value = date
    Gateway2.cell(row=12, column=23).value = date

    openEmpty.save(r"excelFiles\emptyCertificate.xlsx")


def setBench(bench):
    """
    Saves specific data to cells in different sheets of the Empty Certificate based on the provided benches dictionary.

    Args:
        benches (dict): A nested dictionary containing sheet names and their corresponding cell data.
        bench (str): The key to access the specific bench data within the benches dictionary.

    Returns:
        None
    """
    
    openEmpty = load_workbook(r"excelFiles\emptyCertificate.xlsx")

    benches = {
        'WECO4150': {
            '6312': {
                (10, 4): "RX-30-430Xytronic",
                (10, 10): "710194",
                (11, 4): "MD-CC-03",
                (11, 10): "W8777",
                (12, 4): "TRIACTA-6312-JIG-XP-02",
                (12, 10): "ECI-9130JIG-01",
                (13, 4): "MD-POB-06",
                (13, 10): "112020304"
            },
            '6320': {
                (10, 4): "RX-30-430Xytronic",
                (10, 10): "710194",
                (11, 4): "MD-CC-03",
                (11, 10): "W8777",
                (12, 4): "TRIACTA-6312-JIG-XP-02",
                (12, 10): "ECI-9130JIG-01",
                (13, 4): "MD-POB-06",
                (13, 10): "112020304"
            },
            '6303': {
                (10, 4): "RX-30-430Xytronic",
                (10, 10): "710194",
                (11, 4): "MD-CC-03",
                (11, 10): "W8777",
                (12, 4): "TRIACTA-6312-JIG-XP-02",
                (12, 10): "ECI-9130JIG-01",
                (13, 4): "MD-POB-06",
                (13, 10): "112020304"
            },
            '6312-3P': {
                (10, 4): "RX-30-430Xytronic",
                (10, 10): "710194",
                (11, 4): "MD-CC-03",
                (11, 10): "W8777",
                (12, 4): "TRIACTA-6312-JIG-XP-02",
                (12, 10): "ECI-9130JIG-01",
                (13, 4): "MD-POB-06",
                (13, 10): "112020304"
            }
        },
        'WECO2350': {
            '6312': {
                (10, 4): "RM-20-102",
                (10, 10): "206351",
                (11, 4): "MD-CC-02",
                (11, 10): "6500",
                (12, 4): "TRIACTA-6320-JIG-3P-01",
                (12, 10): "ECI-9130JIG-01",
                (13, 4): "MD-POB-03",
                (13, 10): "112020301"
            },
            '6320': {
                (10, 4): "RM-20-102",
                (10, 10): "206351",
                (11, 4): "MD-CC-02",
                (11, 10): "6500",
                (12, 4): "TRIACTA-6320-JIG-3P-01",
                (12, 10): "ECI-9130JIG-01",
                (13, 4): "MD-POB-03",
                (13, 10): "112020301"
            },
            '6303': {
                (10, 4): "RM-20-102",
                (10, 10): "206351",
                (11, 4): "MD-CC-02",
                (11, 10): "6500",
                (12, 4): "TRIACTA-6320-JIG-3P-01",
                (12, 10): "ECI-9130JIG-01",
                (13, 4): "MD-POB-03",
                (13, 10): "112020301"
            },
            '6312-3P': {
                (10, 4): "RM-20-102",
                (10, 10): "206351",
                (11, 4): "MD-CC-02",
                (11, 10): "6500",
                (12, 4): "TRIACTA-6320-JIG-3P-01",
                (12, 10): "ECI-9130JIG-01",
                (13, 4): "MD-POB-03",
                (13, 10): "112020301"
            },
            'Gateway1': {
                (10, 4): "RM-20-102",
                (10, 10): "206351",
                (11, 4): "MD-CC-02",
                (11, 10): "6500",
                (12, 4): "TRIACTA-6320-JIG-3P-01",
                (12, 10): "ECI-9130JIG-01",
                (13, 4): "MD-POB-03",
                (13, 10): "112020301"
            },
            'Gateway2': {
                (10, 4): "RM-20-102",
                (10, 10): "206351",
                (11, 4): "MD-CC-02",
                (11, 10): "6500",
                (12, 4): "TRIACTA-6320-JIG-3P-01",
                (12, 10): "ECI-9130JIG-01",
                (13, 4): "MD-POB-03",
                (13, 10): "112020301"
            },
        }
    }
    
    benchData = benches.get(bench, {}) 
    
    for sheet, cell_data in benchData.items():
        sheet = openEmpty[sheet]
        for (row, column), value in cell_data.items():
            sheet.cell(row=row, column=column).value = value

    openEmpty.save(r"excelFiles\emptyCertificate.xlsx")


def openCertTemp():
    """
    Opens an Excel file named 'emptyCertificate.xlsx' located in the 'excelFiles' directory.

    Returns:
        None
    """
    filePath = r"excelFiles\emptyCertificate.xlsx"

    if os.path.exists(filePath):
        os.startfile(filePath)
        print("Opened empty certificate")
    
    else:
        print("Unable to open certificate")


def openSealinglog():
    """
    Opens an Excel file named 'SealedTest.xlsm' located in the 'excelFiles' directory.

    Returns:
        None
    """
    filePath = r"excelFiles\SealingLog.xlsx"

    if os.path.exists(filePath):
        os.startfile(filePath)
        print("Opened empty certificate")
    
    else:
        print("Unable to open certificate") 


def openMeterError(bench):
    """
    Opens an Excel file containing meter errors based on the specified bench.

    Args:
        bench (str): The bench identifier ('WECO2350' or 'WECO4150').

    Returns:
        None
    """

    if bench == 'WECO2350':
        filePath = r"excelFiles\WECO2350_CertErrors.xlsx"

    elif bench == 'WECO4150':
        filePath = r"excelFiles\WECO4150_CertErrors.xlsx"

    if os.path.exists(filePath):
        os.startfile(filePath)
        print("Opened empty certificate")
    
    else:
        print("Unable to open certificate") 


def openFeedback():
    url = ""

    webbrowser.open(url)

def openLoop():
    url = ""

    webbrowser.open(url)