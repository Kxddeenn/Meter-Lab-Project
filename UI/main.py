import customtkinter as ctk
from tkinter import filedialog
from PIL import Image
import os
from generateCert import cert6303, cert6312, cert6320, certGateway
from components.readXML import findProduct
from openpyxl import load_workbook
from components.exportSealing import exportSealing
import webbrowser
import random

productType = None
customerType = None
meterType = None
badgeType = None
firmwareType = None
XMLfilePath = ""
CSVfilepath = ""
name = None
address = None
regNumber = None



def uiMain(mainScreen):
    """
    Sets up the UI elements for the Main tab.

    Parameters:
    - mainScreen: The main screen (CTkTabview tab) where UI elements will be placed.
    """

    # Global variables for UI elements
    global productType, customerType, meterType, badgeType, firmwareType
    global name, address, regNumber, voltage, serialNumber, viewCert
    global XMLfilePath, CSVfilepath, generateButton, xmlData, rawData, exportButton
    global additionalLabel4, additionalLabel5, additionalLabel6

    def toggleVoltage(value):
        global bench 
        """
        Toggle voltage options based on selected product type.

        Parameters:
        - value: The selected value from productType combobox.
        """

        emptyCert = load_workbook(r"excelFiles\emptyCertificate.xlsx")
        sheet = emptyCert.active

        additionalLabel6.place_forget()
        generateButton.place(x=200, y=400)

        serialNumber.place_forget()
        additionalLabel5.place_forget()

        if sheet.cell(row=10, column=4).value == "RX-30-430Xytronic":
            bench = "WECO4150"

        elif sheet.cell(row=10, column=4).value == "RM-20-102": 
            bench = "WECO2350"

        if bench == "WECO4150":
            if value == "6312" or value == "6303":
                additionalLabel4.configure(state='normal', text="Voltage:")
                additionalLabel4.place(x=295, y=50)

                voltage.configure(width=80, state='normal', values=['120', '240', '277', '347', '480', '600'])
                voltage.place(x=375, y=50)

            else:
                additionalLabel4.configure(state='disabled')
                additionalLabel4.place_forget()

                voltage.configure(width=80, state='normal')
                voltage.place_forget()

        elif bench == "WECO2350":
            if value == "6312" or value == "6303" or value == "Gateway" or value =='6312 (1-12)' or value == '6312 (13-24)':
                additionalLabel4.configure(state='normal', text="Voltage:")
                additionalLabel4.place(x=295, y=50)

                voltage.configure(width=80, state='normal', values=['120', '347', '600'])
                voltage.place(x=375, y=50)

            else:
                additionalLabel4.configure(state='disabled')
                additionalLabel4.place_forget()

                voltage.configure(width=80, state='normal')
                voltage.place_forget()  

             

    def toggleSerialNumb(value):
        """
        Toggle serial number entry based on product type and selected voltage.

        Parameters:
        - value: The selected value from voltage combobox.
        """
        if productType.get() == "6312":
            if voltage.get() == "347" or voltage.get() == "600":
                additionalLabel5.configure(state='normal', text="PT Serial #:")
                additionalLabel5.place(x=295, y=90)

                serialNumber.configure(state='normal', width=115)
                serialNumber.place(x=375, y=90)

            else:
                additionalLabel5.configure(state='disabled', text="PT Serial #:")
                additionalLabel5.place_forget()

                serialNumber.configure(state='disabled', width=115)
                serialNumber.place_forget()    
                
        
    def toggleOther(value):
        """
        Toggle additional fields based on customer type selection.

        Parameters:
        - value: The selected value from customerType combobox.
        """
        if value == "Other":

            additionalLabel1.configure(state='normal', text="Customer Name:")
            additionalLabel1.place(x=295, y=130)

            additionalLabel2.configure(text='Address:', font=('Arial', 13), text_color="white")
            additionalLabel2.place(x=295, y=170)

            additionalLabel3.configure(text='Regulation #:', font=('Arial', 13), text_color="white")
            additionalLabel3.place(x=295, y=210)

            name.configure(width=115, state='normal')
            name.place(x=400, y=130)

            address.configure(width=115, state='normal')
            address.place(x=400, y=170)

            regNumber.configure(width=115, state='normal')
            regNumber.place(x=400, y=210) 

        else:
            additionalLabel1.configure(state='disabled')
            additionalLabel1.place_forget()

            additionalLabel2.configure(state='disabled')
            additionalLabel2.place_forget()

            additionalLabel3.configure(state='disabled')
            additionalLabel3.place_forget()

            name.configure(state='disabled')
            name.place_forget()

            address.configure(state='disabled')
            address.place_forget()

            regNumber.configure(state='disabled')
            regNumber.place_forget()
            

    # UI Design

    lineColor = ctk.CTkImage(
    light_image=Image.open(r"images\linecolor.png"),
    dark_image=Image.open(r"images\linecolor.png"),
    size=(150,1),
    )
    lineColorPic = ctk.CTkLabel(mainScreen, image=lineColor, text="", bg_color="#2B2B2B")
    lineColorPic.place(x=0,y=20)

    lineColor2 = ctk.CTkImage(
    light_image=Image.open(r"images\linecolor.png"),
    dark_image=Image.open(r"images\linecolor.png"),
    size=(550,2),
    )
    lineColorPic2 = ctk.CTkLabel(mainScreen, image=lineColor2, text="", bg_color="#2B2B2B")
    lineColorPic2.place(x=0,y=245)

    darkbackground = ctk.CTkImage(
    light_image=Image.open(r"images\backgroundcolor.png"),
    dark_image=Image.open(r"images\backgroundcolor.png"),
    size=(550,250),
    )
    darkbackgroundl = ctk.CTkLabel(mainScreen, image=darkbackground, text="")
    darkbackgroundl.place(x=0,y=261)

    title = ctk.CTkLabel(
        mainScreen,
        text="Required Data",
        font=("Arial", 20, "bold"),
        text_color="white",
    )
    title.place(x=0, y=5)

    logo = ctk.CTkImage(
        light_image=Image.open(r"UI\images\metergyicon.ico"),
        dark_image=Image.open(r"UI\images\metergyicon.ico"),
        size=(90, 90),
    )
    imageLogo = ctk.CTkLabel(mainScreen, image=logo, text="")
    imageLogo.place(x=410, y=385)

    add1 = ctk.CTkLabel(mainScreen, text="➤", text_color="white", font=("Arial", 20))
    add1.place(x=0, y=50)

    productText = ctk.CTkLabel(
        mainScreen, text="Product Type:", font=("Arial", 13), text_color="white"
    )
    productText.place(x=30, y=50)
    
    productType = ctk.CTkComboBox(mainScreen, values=["6320", "6312", "6303"], command=lambda value: toggleVoltage(productType.get()))
    productType.set("")
    productType.place(x=135, y=50)

    add2 = ctk.CTkLabel(mainScreen, text="➤", text_color="white", font=("Arial", 20))
    add2.place(x=0, y=90)

    customerText = ctk.CTkLabel(
        mainScreen, text="Customer/Owner:", font=("Arial", 13), text_color="white"
    )
    customerText.place(x=30, y=90)

    customerType = ctk.CTkComboBox(mainScreen, values=["Metergy", "Other"], command=lambda value: toggleOther(value))
    customerType.set("Metergy")
    customerType.place(x=135, y=90)

    add3 = ctk.CTkLabel(mainScreen, text="➤", text_color="white", font=("Arial", 20))
    add3.place(x=0, y=130)

    meterText = ctk.CTkLabel(
        mainScreen, text="Verification:", font=("Arial", 13), text_color="white"
    )
    meterText.place(x=30, y=130)

    meterType = ctk.CTkComboBox(mainScreen, values=["Verified", "Re-verified"])
    meterType.set("")
    meterType.place(x=135, y=130)

    add4 = ctk.CTkLabel(mainScreen, text="➤", text_color="white", font=("Arial", 20))
    add4.place(x=0, y=210)

    badgeText = ctk.CTkLabel(
        mainScreen, text="Badge Number:", font=("Arial", 13), text_color="white"
    )
    badgeText.place(x=30, y=210)

    badgeType = ctk.CTkEntry(mainScreen, width=142)
    badgeType.place(x=135, y=210)

    add5 = ctk.CTkLabel(mainScreen, text="➤", text_color="white", font=("Arial", 20))
    add5.place(x=0, y=170)

    firmwareText = ctk.CTkLabel(
        mainScreen, text="Firmware:", font=("Arial", 13), text_color="white"
    )
    firmwareText.place(x=30, y=170)

    firmwareType = ctk.CTkComboBox(mainScreen, values=["2.08", "1.12"])
    firmwareType.set("")
    firmwareType.place(x=135, y=170)


    # Submit files section

    title2 = ctk.CTkLabel(
        mainScreen, text="Submit Files", font=("Arial", 20, "bold"), text_color="white", bg_color="#061E40"
    )
    title2.place(x=210, y=275)

    rawData = ctk.CTkButton(
        mainScreen, text="Raw Data File (.csv)", fg_color="#061E40", command=uploadCSV, border_color="white", border_width=2, font=("Calibri", 15)
    )
    rawData.place(x=300, y=335)

    xmlData = ctk.CTkButton(
        mainScreen, text="XML or Json File", fg_color="#061E40", command=uploadXML, border_color="white", border_width=2, font=("Calibri", 15)
    )
    xmlData.place(x=100, y=335)

    generateButton = ctk.CTkButton(
        mainScreen,
        text="Generate Certificate",
        width=20,
        height=30,
        state="disabled",
        fg_color="#061E40",
        border_color="white",
        font=("Calibri", 15),
        border_width=2,
        command=gatherInputs,
    )
    generateButton.place(x=200, y=400)


    # If Other is selected for Customer, show these labels
    additionalLabel1 = ctk.CTkLabel(mainScreen, font=('Arial', 13), text_color="white", state='disabled')
   
    additionalLabel2 = ctk.CTkLabel(mainScreen, text='Address:', font=('Arial', 13), text_color="white")

    additionalLabel3 = ctk.CTkLabel(mainScreen, text='Regulation #:', font=('Arial', 13), text_color="white")

    name = ctk.CTkEntry(mainScreen, width=115)

    address = ctk.CTkEntry(mainScreen, width=115)

    regNumber = ctk.CTkEntry(mainScreen, width=115)

    # If 6312 is shown, show these labels

    additionalLabel4 = ctk.CTkLabel(mainScreen, font=('Arial', 13), text_color="white", state='disabled')
   
    additionalLabel5 = ctk.CTkLabel(mainScreen, text='Address:', font=('Arial', 13), text_color="white")

    voltage = ctk.CTkComboBox(mainScreen, values=["120", "240", "277", "347", "480", "600"], command=lambda value: toggleSerialNumb(value))

    serialNumber = ctk.CTkEntry(mainScreen, width=115)

    
    # Open Certificate Button

    viewCert = ctk.CTkButton(mainScreen, text="View Certificate", command=openCertificate, width=20, height=30,  font=('Arial', 15), fg_color="#061E40", border_color="white", border_width=2)


    # Export Button 

    exportButton = ctk.CTkButton(mainScreen, text="Export", command=refreshUI, width=70, height=30, fg_color="#061E40", border_color="white", border_width=2)
    additionalLabel6 = ctk.CTkLabel(mainScreen, text="Successfully Exported", font=("Calibri", 20), bg_color="#061E40")

    
def gatherInputs():
    """
    Gather user inputs from UI elements and validate them before generating certificates.
    """
    global \
        productType, \
        customerType, \
        meterType, \
        badgeType, \
        firmwareType, \
        name, \
        address, \
        regNumber, \
        XMLfilePath, \
        CSVfilepath, \
        voltage, \
        serialNumber, \
        viewCert, \
        bench

    productSelection = productType.get()
    customerSelection = customerType.get()
    meterSelection = meterType.get()
    badgeSelection = badgeType.get()
    firmwareSelection = firmwareType.get()
    nameSelection = name.get()
    addressSelection = address.get()
    regNumberSelection = regNumber.get()
    voltageSelection = voltage.get()
    serialNumberSelection = serialNumber.get()

    if(productSelection in ['6320','6312', '6303', 'Gateway', "6312 (1-12)", "6312 (13-24)"] and
       customerSelection in ['Metergy', 'Other'] and
       meterSelection in ["Verified", "Re-verified"] and
       badgeSelection != "" and
       firmwareSelection != ""
       ):
        if XMLfilePath.endswith('.xml'):
            productID = findProduct(XMLfilePath, productSelection)
        if customerSelection == 'Metergy':
            if productSelection == '6320':
                cert6320(productID, 'Metergy', "", "", meterSelection, XMLfilePath, CSVfilepath, badgeSelection, firmwareSelection)
                viewCert.place(x=210, y= 400)

            elif productSelection == '6312':
                cert6312(voltageSelection, productID, "", 'Metergy', "", "", meterSelection, XMLfilePath, CSVfilepath, badgeSelection, firmwareSelection, serialNumberSelection)
                viewCert.place(x=210, y=400)

            elif productSelection == '6303':
                cert6303(voltageSelection, productID, '', '', 'Metergy', meterSelection, XMLfilePath, CSVfilepath, badgeSelection, firmwareSelection)
                viewCert.place(x=210, y=400)
            
            elif productSelection == 'Gateway':
                certGateway(voltageSelection, 'Metergy', "", "", meterSelection, XMLfilePath, CSVfilepath, badgeSelection)
                viewCert.place(x=210, y=400)

            elif productSelection == '6312 (1-12)' or productSelection == '6312 (13-24)':
                cert6312(voltageSelection, "", productSelection, "Metergy", "", "", meterSelection, XMLfilePath, CSVfilepath, badgeSelection, firmwareSelection, serialNumberSelection)
                viewCert.place(x=210, y=400)

        elif customerSelection == "Other":
            if productSelection == '6320':
                cert6320(productID, nameSelection, addressSelection, regNumberSelection, meterSelection, XMLfilePath, CSVfilepath, badgeSelection, firmwareSelection)
                viewCert.place(x=210, y= 400)

            elif productSelection == '6312':
                cert6312(voltageSelection, productID, "", nameSelection, addressSelection, regNumberSelection, meterSelection, XMLfilePath, CSVfilepath, badgeSelection, firmwareSelection, serialNumberSelection)
                viewCert.place(x=210, y=400)

            elif productSelection == '6303':
                cert6303(voltageSelection, productID, addressSelection, regNumberSelection, nameSelection, meterSelection, XMLfilePath, CSVfilepath, badgeSelection, firmwareSelection)
                viewCert.place(x=210, y=400)

            elif productSelection == 'Gateway':
                if bench == 'WECO2350':
                    certGateway(voltageSelection, productID, addressSelection, regNumberSelection, meterSelection, XMLfilePath, CSVfilepath, badgeSelection)
                    viewCert.place(x=210, y=400)
                else: 
                    print(f"Gateway unsupported on {bench}.")

            elif productSelection == '6312 (1-12)' or productSelection == '6312 (13-24)':
                cert6312(voltageSelection, "", productSelection, nameSelection, addressSelection, regNumberSelection ,meterSelection, XMLfilePath, CSVfilepath, badgeSelection, firmwareSelection, serialNumberSelection)
                viewCert.place(x=210, y=400)
        
        generateButton.place_forget()

    else:
        generateButton.configure(border_color="red")
    
# Any additional Events or functions

def uploadXML():
    """
        Allows user to submit their XML file, stores XML path
    """
    global XMLfilePath, xmlData 

    XMLfilePath = filedialog.askopenfilename()

    if XMLfilePath and XMLfilePath.endswith(".xml"):
        xmlData.configure(
            text="XML File Selected", font=("Calibri", 15, "bold"), border_color="green", border_width=3
        )

    elif XMLfilePath and XMLfilePath.endswith('.json'):
        xmlData.configure(
            text="JSON File Selected", font=("Calibri", 15, "bold"), border_color="green", border_width=3
        )
    else:
        xmlData.configure(text="Please Select Proper File", border_color="red", border_width=3)

    updateGenerate()

def uploadCSV():
    """
        Allows user to submit their CSV file, stores CSV path
    """
    global CSVfilepath, rawData, XMLfilePath, generateButton

    CSVfilepath = filedialog.askopenfilename()

    if CSVfilepath and CSVfilepath.endswith(".csv"):
        rawData.configure(
            text="CSV File Selected", font=("Calibri", 15, "bold"), border_color="green", border_width=3
        )

    else:
        rawData.configure(text="Select CSV File (.csv)", border_color="red", border_width=3 )

    updateGenerate()

def updateGenerate():
    """
        Allows user to submit to access the generate button for certificates. (Conditions must be met)
    """
    global generateButton, CSVfilepath, XMLfilePath

    if CSVfilepath and XMLfilePath:
        generateButton.configure(state="normal", font=("Calibri", 15, "bold" ))
    else:
        generateButton.configure(state="disabled")

def openCertificate():
    """
        Allows user to see the certificate that was just created
    """
    filePath = r"outputs\modifiedCert.xlsx"

    if os.path.exists(filePath):
        os.startfile(filePath)
        print("Opened certificate")
    
    else:
        print("Unable to open certificate")

    exportButton.place(x=235, y=450)

def refreshUI():
    productType.set("")
    customerType.set("")
    meterType.set("")
    badgeType.delete(0, ctk.END)
    firmwareType.set("")
    voltage.place_forget()
    serialNumber.place_forget()
    rawData.configure(border_color="white", border_width=2, font=("Calibri", 15, "normal"), text="Raw Data File (.csv)")
    xmlData.configure(border_color="white", border_width=2, font=("Calibri", 15, "normal"), text="XML or Json File")
    generateButton.configure(state='disabled', font=("Calibri", 15, "normal"))
    viewCert.place_forget()
    exportButton.place_forget()
    additionalLabel4.place_forget()
    additionalLabel5.place_forget()
    additionalLabel6.place(x=180, y=400)

    exportSealing()

    # For fun, don't click the link ;)
    randomizer = random.randint(1,100)
    url = "https://www.youtube.com/watch?v=pKskW7wJ0v0"
    
    if randomizer == 50:
        webbrowser.open(url)

def checkBench():
    emptyCert = load_workbook(r"excelFiles\emptyCertificate.xlsx")
    sheet = emptyCert['6312']

    if sheet.cell(row=10, column=4).value == "RX-30-430Xytronic":
        bench2 = "WECO4150"

    elif sheet.cell(row=10, column=4).value == "RM-20-102": 
        bench2 = "WECO2350" 

    if bench2 == 'WECO2350':
        productType.configure(values=["6320", "6312", "6303", "Gateway", "6312 (1-12)", "6312 (13-24)"])

    elif bench2 == 'WECO4150':
        productType.configure(values=["6320", "6312", "6303"])